#!/usr/bin/env python3
"""
grimorium-from-arp.py
Build a Grimorium config from a UniFi gateway's LIVE neighbor (ARP) table,
not from static reservations. Catches dynamic clients too; names them from
the gateway's DHCP leases; optionally re-applies the category sigils from your
reservation spreadsheet by MAC.

HOW IT GETS THE DATA
  One SSH call to the gateway dumps the kernel neighbor table plus the dnsmasq
  lease file. SSH must be enabled on the gateway (UniFi: Settings -> System ->
  Device SSH Authentication). This script never handles your password — it
  shells out to your own `ssh`, so key auth / ssh-agent / an ssh_config host
  alias all work, and a password prompt (if any) happens in your terminal.

  Remote command run (read-only, changes nothing):
    ip neigh show            (falls back to /proc/net/arp)
    cat <first dnsmasq lease file found>

MODES
  Live:     python3 grimorium-from-arp.py --host 192.168.1.1 --user root
  Offline:  ssh root@192.168.1.1 'ip neigh show; echo ===LEASES===; \
              cat /run/dnsmasq.leases' > dump.txt     # capture however you like
            python3 grimorium-from-arp.py --from-file dump.txt
  The offline mode also accepts a raw `ip neigh` dump with no ===LEASES=== marker.

OPTIONS
  --tags-from RES.xlsx   re-apply MGMT/IoT/VINTAGE/... sigils to live devices by MAC
  --reachable-only       keep only REACHABLE/PERMANENT neighbors (drop STALE/DELAY/PROBE)
  --user / --host        SSH target (default root@192.168.1.1)

NEIGHBOR STATES
  Kept by default: REACHABLE, STALE, DELAY, PROBE, PERMANENT (all "seen recently").
  Always dropped: FAILED, INCOMPLETE, NOARP, and entries with no MAC.

ADDRESS FILTERING
  Only RFC1918 / private addresses are kept. Public WAN-side neighbors that the
  gateway tracks (ISP upstream gateway, upstream DNS like 1.1.1.1) are dropped —
  they don't belong on a LAN dashboard and would otherwise leak your ISP/region.
"""

import argparse
import ipaddress
import json
import re
import subprocess
import sys

LEASE_MARK = "===LEASES==="
KEEP_STATES = {"REACHABLE", "STALE", "DELAY", "PROBE", "PERMANENT"}
STRICT_STATES = {"REACHABLE", "PERMANENT"}
IPV4_RE = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")
MAC_RE = re.compile(r"^[0-9a-f]{2}(:[0-9a-f]{2}){5}$")

# Candidate dnsmasq lease paths across USG / UDM / UCG / UniFi OS versions.
LEASE_PATHS = [
    "/run/dnsmasq.leases", "/var/run/dnsmasq.leases",
    "/var/lib/misc/dnsmasq.leases", "/var/lib/dhcp/dnsmasq.leases",
    "/data/udapi-config/dnsmasq.lease", "/data/unifi-core/config/dnsmasq.leases",
]

REMOTE_CMD = (
    "ip neigh show 2>/dev/null || cat /proc/net/arp; "
    f"echo {LEASE_MARK}; "
    "for f in " + " ".join(LEASE_PATHS) + "; do "
    "[ -f \"$f\" ] && cat \"$f\" && break; done"
)

TAG_STYLE = {
    "MGMT": ("\u2699", "#7c9ed8"), "IoT": ("\u26a1", "#ffcf3f"),
    "GENERAL": ("\u2726", "#8ee066"), "VINTAGE": ("\u2638", "#d18b1d"),
    "ESXi VMs": ("\u25a3", "#c87ad1"), "DHCP": ("\u2042", "#7adcc7"),
}
SITE_STYLE = {
    "Home LAN": ("\u2302", "#8ee066"), "VMF LAN": ("\u2349", "#b8521c"),
    "Pentagon City LAN": ("\u2316", "#d87a7a"), "L2TP VPN": ("\u26bf", "#7c9ed8"),
}
OCTET_SITE = {0: "VMF LAN", 1: "Home LAN", 2: "L2TP VPN", 3: "Pentagon City LAN"}


def norm_mac(s):
    """Lowercase, trim, unify separators. '' if not a MAC."""
    if not s:
        return ""
    s = str(s).strip().lower().replace("-", ":")
    return s if MAC_RE.match(s) else ""


def link_for(host, ip):
    """Best-guess single liveness link (opaque https probe) from the device name."""
    h = (host or "").lower()
    if any(k in h for k in ("synology", "xpenology", "nas")):
        return {"name": "reachable", "probe": "https", "target": f"https://{ip}:5001/"}
    if "idrac" in h or h.startswith("esxi") or "vsphere" in h or h.startswith(("usg", "usw", "uap", "us8", "us24", "udm", "ucg")):
        return {"name": "reachable", "probe": "https", "target": f"https://{ip}/"}
    return {"name": "reachable", "probe": "https", "target": f"http://{ip}/"}


def get_dump(args):
    if args.from_file:
        with open(args.from_file, encoding="utf-8", errors="replace") as f:
            return f.read()
    target = f"{args.user}@{args.host}"
    try:
        out = subprocess.run(
            ["ssh", "-o", "BatchMode=no", "-o", "ConnectTimeout=8", target, REMOTE_CMD],
            capture_output=True, text=True, timeout=40,
        )
    except FileNotFoundError:
        sys.exit("ssh not found on PATH")
    except subprocess.TimeoutExpired:
        sys.exit(f"ssh to {target} timed out")
    if out.returncode != 0:
        sys.exit(f"ssh failed ({out.returncode}): {out.stderr.strip()}")
    return out.stdout


def split_dump(text):
    if LEASE_MARK in text:
        neigh, leases = text.split(LEASE_MARK, 1)
    else:
        neigh, leases = text, ""
    return neigh, leases


def parse_neigh(text, strict):
    """Yield (ip, mac) for usable IPv4 neighbors. Handles `ip neigh` and /proc/net/arp."""
    allowed = STRICT_STATES if strict else KEEP_STATES
    out = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.lower().startswith("ip address"):  # /proc header
            continue
        toks = line.split()
        # /proc/net/arp: IP HWtype Flags HWaddr Mask Device
        if len(toks) == 6 and IPV4_RE.match(toks[0]) and norm_mac(toks[3]):
            if toks[2] != "0x0" and toks[3] != "00:00:00:00:00:00":
                out.append((toks[0], norm_mac(toks[3])))
            continue
        # ip neigh:  IP dev X lladdr MAC STATE   (state token is the trailing word)
        if not IPV4_RE.match(toks[0]):
            continue
        mac = ""
        if "lladdr" in toks:
            i = toks.index("lladdr")
            if i + 1 < len(toks):
                mac = norm_mac(toks[i + 1])
        if not mac:
            continue
        last = toks[-1].upper()
        state = last if last.isalpha() else None       # None => no explicit state
        if state is None or state in allowed:           # drop only explicit FAILED/INCOMPLETE/NOARP
            out.append((toks[0], mac))
    return out


def parse_leases(text):
    """dnsmasq leases: '<expiry> <mac> <ip> <name> <clientid>'. Return mac->name, ip->name."""
    by_mac, by_ip = {}, {}
    for line in text.splitlines():
        toks = line.split()
        if len(toks) < 4:
            continue
        mac, ip, name = norm_mac(toks[1]), toks[2], toks[3]
        if name in ("*", ""):
            continue
        if mac:
            by_mac[mac] = name
        if IPV4_RE.match(ip):
            by_ip[ip] = name
    return by_mac, by_ip


def macmap_from_xlsx(path):
    """From the reservation sheet: MAC -> category tag (col G, forward-filled) and
    MAC -> curated host name (col C). Returns (mac_tag, mac_name)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb["Current LAN"] if "Current LAN" in wb.sheetnames else wb[wb.sheetnames[0]]
    mac_tag, mac_name, current = {}, {}, None
    for r in range(2, ws.max_row + 1):
        g = ws.cell(r, 7).value
        if g not in (None, ""):
            current = str(g).strip()
        mac = norm_mac(ws.cell(r, 4).value)
        host = ws.cell(r, 3).value
        if mac:
            if current:
                mac_tag[mac] = current
            if host not in (None, ""):
                mac_name[mac] = str(host).strip()
        # side tables (cols H..M): pick up MAC -> name so off-subnet boxes get names too
        for c in range(8, 14):
            sm = norm_mac(ws.cell(r, c).value)
            if sm and sm not in mac_name:
                # name is the nearest non-IP, non-MAC token on the row
                for cc in range(8, 14):
                    tok = ws.cell(r, cc).value
                    s = str(tok).strip() if tok is not None else ""
                    if s and not MAC_RE.match(s.lower().replace("-", ":")) \
                       and not IPV4_RE.match(s) and "/" not in s \
                       and not s.startswith("dhcp-host=") and not s.endswith(("LAN", "VPN")):
                        mac_name[sm] = s
                        break
    return mac_tag, mac_name


def build(args):
    neigh_txt, lease_txt = split_dump(get_dump(args))
    neighbors = parse_neigh(neigh_txt, args.reachable_only)
    by_mac, by_ip = parse_leases(lease_txt)
    mac_tag, mac_name = macmap_from_xlsx(args.tags_from) if args.tags_from else ({}, {})

    # De-dupe by IP, first wins. Drop non-private (WAN-side) neighbors such as
    # the ISP gateway or upstream DNS — they show up in the gateway's neighbor
    # table but have no place on a LAN dashboard (and leak your ISP/region).
    seen, hosts, dropped_public = set(), [], 0
    for ip, mac in neighbors:
        try:
            if not ipaddress.ip_address(ip).is_private:
                dropped_public += 1
                continue
        except ValueError:
            continue
        if ip in seen:
            continue
        seen.add(ip)
        # name priority: curated reservation name > DHCP lease name > IP
        name = mac_name.get(mac) or by_mac.get(mac) or by_ip.get(ip) or ip
        site = OCTET_SITE.get(int(ip.split(".")[2]))
        hosts.append({"ip": ip, "mac": mac, "name": name, "tag": mac_tag.get(mac), "site": site})

    classifiers, cls_id = [], {}
    def ensure(name, style):
        if not name or name in cls_id:
            return
        glyph, tint = style.get(name, ("\u2726", "#8ee066"))
        cid = "cls-" + re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
        cls_id[name] = cid
        classifiers.append({"id": cid, "name": name, "glyph": glyph, "tint": tint})
    for h in hosts:
        ensure(h["tag"], TAG_STYLE)
    for h in hosts:
        ensure(h["site"], SITE_STYLE)

    chains = []
    for h in hosts:
        cids = [cls_id[h["tag"]]] if h["tag"] in cls_id else []
        if h["site"] in cls_id:
            cids.append(cls_id[h["site"]])
        chains.append({
            "name": h["name"], "address": h["ip"], "haltOnFail": False,
            "classifierIds": cids, "links": [link_for(h["name"], h["ip"])],
        })

    return {
        "chains": chains, "classifiers": classifiers, "positions": {},
        "groupByTag": True, "timeoutMs": 5000, "parallel": 6,
    }, len(chains), dropped_public


def main():
    ap = argparse.ArgumentParser(description="UniFi ARP table -> Grimorium import JSON")
    ap.add_argument("--host", default="192.168.1.1", help="gateway IP/host (default 192.168.1.1)")
    ap.add_argument("--user", default="root", help="SSH user (default root)")
    ap.add_argument("--from-file", help="parse a captured dump instead of SSHing")
    ap.add_argument("--tags-from", help="reservation .xlsx to source category sigils by MAC")
    ap.add_argument("--reachable-only", action="store_true",
                    help="keep only REACHABLE/PERMANENT neighbors")
    args = ap.parse_args()

    config, n, dropped = build(args)
    json.dump(config, sys.stdout, indent=2, ensure_ascii=False)
    sys.stdout.write("\n")
    note = f" ({dropped} public/WAN neighbor{'s' if dropped != 1 else ''} dropped)" if dropped else ""
    print(f"// {n} live hosts -> chains{note}", file=sys.stderr)


if __name__ == "__main__":
    main()
